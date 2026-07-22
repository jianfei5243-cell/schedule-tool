#!/usr/bin/env python3
"""
中框测试排期自动生成脚本
=======================
根据评审时间倒排测试周期，自动计算 LQ/MI 测试起止日期，
验证阶段间不重叠、启动时间在上阶段评审之后，可导出 Excel。

用法:
    python3 generate_schedule.py                          # 使用内置默认配置
    python3 generate_schedule.py --config my_config.json  # 使用 JSON 配置文件
    python3 generate_schedule.py --excel 排期表.xlsx       # 导出到指定 Excel 文件
    python3 generate_schedule.py --help                   # 查看帮助

依赖: Python 3.7+, openpyxl (可选，用于 Excel 导出)
"""

import argparse
import json
import sys
from datetime import date, timedelta
from pathlib import Path

# ============================================================
#  节假日日历
# ============================================================

class HolidayCalendar:
    """管理节假日，支持批量添加和单个查询"""

    def __init__(self):
        self._holidays: set[date] = set()

    def add(self, *dates: date):
        """添加节假日"""
        self._holidays.update(dates)

    def add_range(self, start: date, end: date):
        """添加一段连续的节假日"""
        d = start
        while d <= end:
            self._holidays.add(d)
            d += timedelta(days=1)

    def is_working(self, d: date) -> bool:
        """检查是否为工作日（周一至周五且非节假日）"""
        return d.weekday() < 5 and d not in self._holidays

    @staticmethod
    def weekday_cn(d: date) -> str:
        """返回中文星期"""
        return "一二三四五六日"[d.weekday()]

    @staticmethod
    def date_fmt(d: date) -> str:
        """格式化日期 YYYY/MM/DD"""
        return f"{d.year}/{d.month:02d}/{d.day:02d}"


# ============================================================
#  核心排期引擎
# ============================================================

class ScheduleEngine:
    """根据评审日期倒排测试排期"""

    def __init__(self, calendar: HolidayCalendar):
        self.cal = calendar

    # ── 日期工具 ──

    def next_working(self, anchor: date) -> date:
        """第一个 >= anchor 的工作日"""
        d = anchor
        while not self.cal.is_working(d):
            d += timedelta(days=1)
        return d

    def prev_working(self, anchor: date) -> date:
        """最后一个 < anchor 的工作日"""
        d = anchor - timedelta(days=1)
        while not self.cal.is_working(d):
            d -= timedelta(days=1)
        return d

    def count_back(self, end_date: date, n_days: int) -> date:
        """从 end_date（含）往前数 n 个工作日，返回起始日期"""
        d = end_date
        found = 0
        while found < n_days:
            if self.cal.is_working(d):
                found += 1
            if found < n_days:
                d -= timedelta(days=1)
        return d

    def working_days_between(self, a: date, b: date) -> int:
        """a 到 b（含）之间的工作日数"""
        return sum(1 for i in range((b - a).days + 1)
                   if self.cal.is_working(a + timedelta(i)))

    # ── 单阶段排期 ──

    def schedule_phase(self, review: date, lq_days: int, mi_days: int = 0):
        """
        倒排一个阶段的 LQ / MI 测试周期。

        返回: { "lq_start", "lq_end", "mi_start", "mi_end" }
        """
        result = {}
        if mi_days > 0:
            mi_end  = self.prev_working(review)
            mi_beg  = self.count_back(mi_end, mi_days)
            lq_end  = self.prev_working(mi_beg)
            lq_beg  = self.count_back(lq_end, lq_days)
            result["mi_start"] = mi_beg
            result["mi_end"]   = mi_end
        else:
            lq_end = self.prev_working(review)
            lq_beg = self.count_back(lq_end, lq_days)
        result["lq_start"] = lq_beg
        result["lq_end"]   = lq_end
        return result

    # ── 全流程排期 + 校验 ──

    def schedule_all(self, phases: list[dict],
                     cp_lq: int = 14, st_lq: int = 12, st_mi: int = 7):
        """
        排全部阶段并校验。

        phases: [{"name": "CP0-CP1", "review": "2026/11/16", "has_mi": False}, ...]
        返回: [{
            "name", "review", "has_mi",
            "lq_start", "lq_end", "mi_start", "mi_end",
            "lq_days", "mi_days", "window_begin", "window_end",
            "window_days", "status", "issue"
        }, ...]
        """
        rows = []
        prev_review = None

        for ph in phases:
            R = date.fromisoformat(ph["review"])
            is_st = "ST" in ph["name"]
            has_mi = ph.get("has_mi", False)

            lq_d = st_lq if is_st else cp_lq
            mi_d = st_mi if (is_st and has_mi) else 0

            # 计算可用窗口（上阶段评审后 → 本次评审前）
            if prev_review is None:
                win_begin = None
                win_end = None
                win_days = None
            else:
                win_begin = self.next_working(prev_review + timedelta(days=1))
                win_end   = self.prev_working(R)
                win_days  = self.working_days_between(win_begin, win_end)

            # 倒排
            sched = self.schedule_phase(R, lq_d, mi_d)

            # 校验
            status = "✓"
            issue  = ""
            if prev_review is not None:
                total_needed = lq_d + mi_d
                if win_days < total_needed:
                    status = "❌"
                    issue = f"窗口仅{win_days}工作日，需{total_needed}天"
                elif sched["lq_start"] < win_begin:
                    status = "⚠️"
                    issue = f"倒排LQ起于{self.cal.date_fmt(sched['lq_start'])}，早于窗口起{self.cal.date_fmt(win_begin)}"

            rows.append({
                "name":        ph["name"],
                "review":      R,
                "has_mi":      has_mi,
                "lq_start":    sched["lq_start"],
                "lq_end":      sched["lq_end"],
                "mi_start":    sched.get("mi_start"),
                "mi_end":      sched.get("mi_end"),
                "lq_days":     lq_d,
                "mi_days":     mi_d,
                "window_begin": win_begin,
                "window_end":   win_end,
                "window_days":  win_days,
                "status":       status,
                "issue":        issue,
            })
            prev_review = R

        return rows


# ============================================================
#  输出：控制台
# ============================================================

def print_console(product_name: str, rows: list, cp_lq: int, st_lq: int, st_mi: int):
    """打印排期到控制台"""
    cal = HolidayCalendar()
    print(f"\n{'='*95}")
    print(f"  {product_name}")
    print(f"  CP: LQ={cp_lq}天 | ST: LQ={st_lq}天 + MI={st_mi}天")
    print(f"{'='*95}")
    print(f"{'阶段':<8s}  测试端   开始时间  (周几)  ~  结束时间  (周几)   天数    校验")
    print(f"{'-'*95}")

    for r in rows:
        name = r["name"]
        lqs = cal.date_fmt(r["lq_start"])
        lqe = cal.date_fmt(r["lq_end"])
        rev = cal.date_fmt(r["review"])

        print(f"{name:<8s}  LQ测试  {lqs} ({cal.weekday_cn(r['lq_start'])})  ~  "
              f"{lqe} ({cal.weekday_cn(r['lq_end'])})  {r['lq_days']}天")

        if r["has_mi"]:
            mis = cal.date_fmt(r["mi_start"])
            mie = cal.date_fmt(r["mi_end"])
            print(f"{'':8s}  MI测试  {mis} ({cal.weekday_cn(r['mi_start'])})  ~  "
                  f"{mie} ({cal.weekday_cn(r['mi_end'])})  {r['mi_days']}天")

        check = r["status"]
        if r["issue"]:
            check += " " + r["issue"]
        print(f"{'':8s}  ── 评审 ──  {rev} ({cal.weekday_cn(r['review'])})    {check}")
        print()


# ============================================================
#  输出：Excel
# ============================================================

def export_excel(products: list[dict], path: str):
    """导出到 Excel 文件"""
    try:
        import openpyxl
    except ImportError:
        print("❌ 需要安装 openpyxl: pip3 install openpyxl")
        return

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # 样式
    hdr_font  = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    hdr_fill  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    lq_fill   = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    mi_fill   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    rv_fill   = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    err_fill  = PatternFill(start_color="F4B4C2", end_color="F4B4C2", fill_type="solid")
    title_font= Font(name="微软雅黑", bold=True, size=14, color="2F5496")
    sub_font  = Font(name="微软雅黑", size=10, color="808080")
    norm_font = Font(name="微软雅黑", size=10)
    bold_font = Font(name="微软雅黑", bold=True, size=10)
    border    = Border(
        left=Side("thin","B4C6E7"), right=Side("thin","B4C6E7"),
        top=Side("thin","B4C6E7"),  bottom=Side("thin","B4C6E7"),
    )
    ca = Alignment(horizontal="center", vertical="center", wrap_text=True)

    tab_colors = ["2F5496", "C55A11", "548235", "7030A0", "BF8F00"]

    wb = openpyxl.Workbook()
    cal = HolidayCalendar()

    for idx, prod in enumerate(products):
        if idx == 0:
            ws = wb.active
        else:
            ws = wb.create_sheet()
        ws.title = prod["name"][:31]
        ws.sheet_properties.tabColor = tab_colors[idx % len(tab_colors)]

        # 标题行
        ws.merge_cells("A1:H1")
        ws["A1"] = f"{prod['name']} 测试排期"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:H2")
        ws["A2"] = f"CP: LQ={prod['cp_lq']}天 | ST: LQ={prod['st_lq']}天 + MI={prod['st_mi']}天 | 已剔除周末与节假日"
        ws["A2"].font = sub_font
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20

        # 表头
        headers = ["阶段", "测试端", "开始时间", "开始周几", "结束时间", "结束周几", "天数", "校验"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=4, column=col, value=h)
            c.font = hdr_font; c.fill = hdr_fill; c.alignment = ca; c.border = border
        ws.row_dimensions[4].height = 22
        widths = [10, 8, 14, 8, 14, 8, 8, 22]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64+i) if i<27 else "A"].width = w
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # 数据行
        row = 5
        for r in prod["rows"]:
            name = r["name"]
            rev  = cal.date_fmt(r["review"])

            def write_line(phase_label, test_label, start, s_wd, end, e_wd, days, is_review, is_error):
                fill = rv_fill if is_review else (err_fill if is_error else (lq_fill if "LQ" in str(test_label) else mi_fill))
                font = bold_font if is_review else norm_font
                vals = [
                    phase_label,
                    test_label,
                    start, s_wd, end, e_wd,
                    days,
                    r["status"] + (" " + r["issue"] if r["issue"] else ""),
                ]
                for col, v in enumerate(vals, 1):
                    c = ws.cell(row=row, column=col, value=v)
                    c.font = font; c.fill = fill; c.alignment = ca; c.border = border

            is_err = r["status"] != "✓"
            write_line(name, "LQ测试",
                       cal.date_fmt(r["lq_start"]), cal.weekday_cn(r["lq_start"]),
                       cal.date_fmt(r["lq_end"]), cal.weekday_cn(r["lq_end"]),
                       f"{r['lq_days']}天", False, is_err)
            row += 1

            if r["has_mi"]:
                write_line(name, "MI测试",
                           cal.date_fmt(r["mi_start"]), cal.weekday_cn(r["mi_start"]),
                           cal.date_fmt(r["mi_end"]), cal.weekday_cn(r["mi_end"]),
                           f"{r['mi_days']}天", False, False)
                # Merge phase cell
                ws.merge_cells(start_row=row-1, start_column=1, end_row=row, end_column=1)
                row += 1

            write_line(name, "评审时间", rev, cal.weekday_cn(r["review"]), "", "", "", True, False)
            row += 1

        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = "landscape"
        ws.freeze_panes = "A5"

    wb.save(path)
    print(f"✅ Excel 已保存: {path}")


# ============================================================
#  默认节假日
# ============================================================

def default_holidays() -> HolidayCalendar:
    """返回包含常用中国节假日的日历"""
    cal = HolidayCalendar()
    # 2027 元旦
    cal.add(date(2027, 1, 1))
    # 2027 春节（预估 2/5-2/11）
    cal.add_range(date(2027, 2, 5), date(2027, 2, 11))
    # 2027 清明
    cal.add(date(2027, 4, 5))
    # 2027 劳动节
    cal.add_range(date(2027, 5, 1), date(2027, 5, 5))
    return cal


# ============================================================
#  默认项目配置（三款中框）
# ============================================================

DEFAULT_CONFIG = {
    "cp_lq": 14,   # CP 阶段 LQ 测试工作日
    "st_lq": 12,   # ST 阶段 LQ 测试工作日
    "st_mi": 7,    # ST 阶段 MI 测试工作日
    "products": [
        {
            "name": "铝合金包皮中框",
            "cp_lq": 13,  # 覆盖默认 CP LQ
            "phases": [
                {"name": "CP0-CP1",  "review": "2026-11-16", "has_mi": False},
                {"name": "CP2",      "review": "2026-12-04", "has_mi": False},
                {"name": "ST1",      "review": "2027-01-04", "has_mi": True},
                {"name": "ST2",      "review": "2027-02-03", "has_mi": True},
                {"name": "国际ST3",  "review": "2027-03-10", "has_mi": True},
            ]
        },
        {
            "name": "纳米注塑中框(双模激进)",
            "phases": [
                {"name": "CP0-CP1",  "review": "2026-11-16", "has_mi": False},
                {"name": "CP2",      "review": "2026-12-09", "has_mi": False},
                {"name": "ST1",      "review": "2027-01-20", "has_mi": True},
                {"name": "ST2",      "review": "2027-03-01", "has_mi": True},
                {"name": "国际ST3",  "review": "2027-04-05", "has_mi": True},
            ]
        },
        {
            "name": "纳米注塑中框(正常)",
            "phases": [
                {"name": "CP0-CP1",  "review": "2026-11-11", "has_mi": False},
                {"name": "CP2",      "review": "2026-12-11", "has_mi": False},
                {"name": "ST1",      "review": "2027-01-28", "has_mi": True},
                {"name": "ST2",      "review": "2027-03-23", "has_mi": True},
                {"name": "国际ST3",  "review": "2027-04-30", "has_mi": True},
            ]
        },
    ]
}


# ============================================================
#  主入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="中框测试排期自动生成脚本",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  %(prog)s                             使用内置默认配置
  %(prog)s --config my_config.json     使用自定义 JSON 配置
  %(prog)s --excel 排期表.xlsx          导出到 Excel
  %(prog)s --excel output.xlsx --no-console  仅导出 Excel
        """,
    )
    parser.add_argument("--config", "-c", type=str, help="JSON 配置文件路径")
    parser.add_argument("--excel", "-e", type=str, help="导出 Excel 文件路径")
    parser.add_argument("--no-console", action="store_true", help="不打印控制台输出")
    parser.add_argument("--cp-lq", type=int, help="CP 阶段 LQ 测试天数（覆盖配置）")
    parser.add_argument("--st-lq", type=int, help="ST 阶段 LQ 测试天数（覆盖配置）")
    parser.add_argument("--st-mi", type=int, help="ST 阶段 MI 测试天数（覆盖配置）")
    args = parser.parse_args()

    # 加载配置
    if args.config:
        with open(args.config, "r", encoding="utf-8") as f:
            config = json.load(f)
    else:
        config = DEFAULT_CONFIG

    # 允许命令行覆盖天数
    global_cp = args.cp_lq or config.get("cp_lq", 14)
    global_st_lq = args.st_lq or config.get("st_lq", 12)
    global_st_mi = args.st_mi or config.get("st_mi", 7)

    # 初始化
    cal = default_holidays()
    engine = ScheduleEngine(cal)
    products = []

    for prod in config["products"]:
        cp  = prod.get("cp_lq", global_cp)
        slq = prod.get("st_lq", global_st_lq)
        smi = prod.get("st_mi", global_st_mi)
        rows = engine.schedule_all(prod["phases"], cp_lq=cp, st_lq=slq, st_mi=smi)

        products.append({
            "name":   prod["name"],
            "cp_lq":  cp,
            "st_lq":  slq,
            "st_mi":  smi,
            "rows":   rows,
        })

        # 控制台输出
        if not args.no_console:
            print_console(prod["name"], rows, cp, slq, smi)

    # Excel 导出
    if args.excel:
        export_excel(products, args.excel)

    # 汇总
    total_ok = sum(1 for p in products for r in p["rows"] if r["status"] == "✓")
    total_all = sum(len(p["rows"]) for p in products)
    if not args.no_console and total_all > 0:
        print(f"\n{'='*50}")
        print(f"  汇总: {total_ok}/{total_all} 阶段通过校验")
        if total_ok < total_all:
            print(f"  ⚠️  有 {total_all - total_ok} 个阶段存在排期冲突，请检查")
        print(f"{'='*50}")


if __name__ == "__main__":
    main()
